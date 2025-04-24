from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
from math import sqrt

class ProcessPartA:
    def __init__(self):
        self.data = {
            "drillholes": [],
            "samples": [],
            "extra": [],
        }

    def retrieve_data(self, file_name):
        try:
            workbook = load_workbook(filename=file_name)
            input_data_sheets = ["DRILLHOLES", "SAMPLES"]
            for sheet in input_data_sheets:
                if sheet not in workbook.sheetnames:
                    raise ValueError(f"Input data sheet '{sheet}' is missing.")
                
            self.data["drillholes"] = self.load_sheet(workbook, "DRILLHOLES")
            self.data["samples"] = self.load_sheet(workbook, "SAMPLES")

            if "EXTRA_DH_DATA" in workbook.sheetnames:
                self.data["extra"] = self.load_sheet(workbook, "EXTRA_DH_DATA")

        except InvalidFileException:
            print(f"Error: '{file_name}' is not valid.")
        except FileNotFoundError:
            print(f"Error: File '{file_name}' does not exist.")
        except KeyError as e:
            print(f"Missing expected sheet: {e}")
        except Exception as e:
            print(f"Unexpected error: {e}")

    def load_sheet(self, workbook, sheet_name):
        try:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            if not headers or None in headers:
                raise ValueError(f"'{sheet_name}' is misisng headers or having empty columns.")
            
            sheet_data = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
            if not sheet_data:
                print(f"Warning: '{sheet_name} sheet is empty.'")
            return sheet_data
        
        except KeyError:
            raise ValueError(f"'{sheet_name}' sheet does not exist.")

    def calculate_total_drilled_length(self):
        return sum(hole.get("Length (m)", 0) for hole in self.data["drillholes"])

    def calculate_average_au_grade(self):
        total_grade = 0
        count = 0
        for sample in self.data["samples"]:
            au = sample.get("Au")
            if isinstance(au, (int, float)):
                    total_grade += au
                    count += 1
        return total_grade / count if count else 0.0

    def get_total_drilled_by_company(self):
        company_totals = {}

        for record in self.data["extra"]:
            if record["Item"] == "COMPANY":
                # Get company name
                hole_name = record["Name"]
                company = record["Value"]

                # Find matching drillhole
                hole = next((h for h in self.data["drillholes"] if h["Name"] == hole_name), None)
                if hole:
                    # Add length to total drilled meters of company
                    meters = hole.get("Length (m)", 0)
                    company_totals[company] = company_totals.get(company, 0) + meters

        return company_totals

    def get_daily_drilled_meters(self):
        day_total = {}
        for data in self.data["extra"]:
            if data["Item"] == "DATE":
                hole_name = data["Name"]
                data_date = data["Value"]
                try:
                    date = datetime.strptime(data_date, "%d-%m-%Y").strftime("%d-%m-%Y")
                except Exception:
                    print(f"Invalid date format: {data_date}")
                    continue

                hole = next((h for h in self.data["drillholes"] if h["Name"] == hole_name), None)
                if hole:
                    length = hole.get("Length (m)", 0)
                    day_total[date] = day_total.get(date, 0) + length
        return day_total 
    
    def print_distances_to_point(self, x, y):
        print(f"\n* Distances from new hole at ({x}, {y}):\n" + "-" * 40)
        for drillhole in self.data["drillholes"]:
            name = drillhole["Name"]
            dx = drillhole["X"] - x
            dy = drillhole["Y"] - y
            distance = round(sqrt(dx**2 + dy**2), 2)
            print(f"\n{name}: {distance} meters")

    def estimate_augrade_from_nearest(self, x, y, n=4):
        distances = {}
        for drillhole in self.data["drillholes"]:
            name = drillhole["Name"]
            dx = drillhole["X"] - x
            dy = drillhole["Y"] - y
            distances[name] = sqrt(dx**2 + dy**2)

        # Get n closest drillholes
        nearest_drillholes = sorted(distances, key=distances.get)[:n]

        # Collect Au values
        total_au_grade = 0 
        count = 0
        for sample in self.data["samples"]:
            if sample["Name"] in nearest_drillholes:
                au = sample.get("Au")
                if isinstance(au, (int, float)):
                    total_au_grade += au
                    count += 1
        average_au = total_au_grade / count if count else 0.0
        print(f"\n* Estimate Au grade at ({x}, {y}) from the {n} closest holes: {average_au:.2f}")
        return average_au
